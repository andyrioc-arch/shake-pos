"""Exportación de los estados financieros a xlsx (openpyxl)."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from . import posting

MONEY = '"$"#,##0.00'
_TITLE = Font(bold=True, size=14)
_SUB = Font(italic=True, color="666666")
_HEAD_FONT = Font(bold=True, color="FFFFFF")
_HEAD_FILL = PatternFill("solid", fgColor="000000")
_SECTION = Font(bold=True, color="FFFFFF")
_SECTION_FILL = PatternFill("solid", fgColor="1E9AFF")
_TOTAL = Font(bold=True)
_TOTAL_FILL = PatternFill("solid", fgColor="EDEDED")
_RIGHT = Alignment(horizontal="right")


def _titulo(ws, titulo, periodo):
    ws["A1"] = titulo
    ws["A1"].font = _TITLE
    ws["A2"] = f"SHAKE · {periodo}"
    ws["A2"].font = _SUB


def _cabecera(ws, fila, encabezados):
    for col, texto in enumerate(encabezados, 1):
        c = ws.cell(row=fila, column=col, value=texto)
        c.font = _HEAD_FONT
        c.fill = _HEAD_FILL


def _seccion(ws, fila, texto, ncols):
    for col in range(1, ncols + 1):
        c = ws.cell(row=fila, column=col)
        c.fill = _SECTION_FILL
    c = ws.cell(row=fila, column=1, value=texto)
    c.font = _SECTION


def _money(ws, fila, col, valor, negrita=False):
    c = ws.cell(row=fila, column=col, value=float(valor or 0))
    c.number_format = MONEY
    if negrita:
        c.font = _TOTAL
    return c


def _total(ws, fila, etiqueta, valor, col_valor):
    ws.cell(row=fila, column=1, value=etiqueta).font = _TOTAL
    for col in range(1, col_valor + 1):
        ws.cell(row=fila, column=col).fill = _TOTAL_FILL
    _money(ws, fila, col_valor, valor, negrita=True)


def _anchos(ws, anchos):
    for col, ancho in enumerate(anchos, 1):
        ws.column_dimensions[chr(64 + col)].width = ancho


# ── Estado de Resultados ──────────────────────────────────────────────────────
def resultados(ws, anio, mes, periodo):
    d = posting.estado_resultados(anio, mes)
    _titulo(ws, "Estado de Resultados", periodo)
    r = 4
    _seccion(ws, r, "INGRESOS", 2); r += 1
    for it in d["ingresos"]:
        ws.cell(row=r, column=1, value=it["nombre"])
        _money(ws, r, 2, it["monto"]); r += 1
    _total(ws, r, "Total ingresos", d["total_ingresos"], 2); r += 2
    _seccion(ws, r, "COSTO DE VENTAS (FIFO)", 2); r += 1
    for it in d["costo_ventas"]:
        ws.cell(row=r, column=1, value=it["nombre"])
        _money(ws, r, 2, it["monto"]); r += 1
    _total(ws, r, "Utilidad bruta", d["utilidad_bruta"], 2); r += 2
    _seccion(ws, r, "GASTOS OPERATIVOS", 2); r += 1
    for it in d["gastos"]:
        ws.cell(row=r, column=1, value=it["nombre"])
        _money(ws, r, 2, it["monto"]); r += 1
    _total(ws, r, "Total gastos operativos", d["total_gastos"], 2); r += 2
    _total(ws, r, "UTILIDAD NETA DEL PERIODO", d["utilidad"], 2)
    _anchos(ws, [34, 16])
    return "estado_resultados"


# ── Balance General ───────────────────────────────────────────────────────────
def balance(ws, anio, mes, periodo):
    d = posting.balance_general(anio, mes)
    _titulo(ws, "Balance General", periodo)
    r = 4

    def bloque(titulo, items, total_lbl, total_val):
        nonlocal r
        _seccion(ws, r, titulo, 2); r += 1
        for it in items:
            ws.cell(row=r, column=1, value=it["nombre"])
            _money(ws, r, 2, it["monto"]); r += 1
        _total(ws, r, total_lbl, total_val, 2); r += 2

    bloque("ACTIVO", d["activos"], "Total activo", d["total_activo"])
    bloque("PASIVO", d["pasivos"], "Total pasivo", d["total_pasivo"])
    bloque("CAPITAL", d["capital"], "Total capital", d["total_capital"])
    _total(ws, r, "TOTAL PASIVO + CAPITAL", d["total_pasivo_capital"], 2); r += 1
    ws.cell(row=r, column=1,
            value=("✓ El balance cuadra" if d["cuadra"] else "⚠ El balance NO cuadra"))
    _anchos(ws, [34, 16])
    return "balance_general"


# ── Flujo de Efectivo ─────────────────────────────────────────────────────────
def flujo(ws, anio, mes, periodo):
    d = posting.flujo_efectivo(anio, mes)
    _titulo(ws, "Flujo de Efectivo", periodo)
    r = 4
    _total(ws, r, "Saldo inicial", d["saldo_inicial"], 3); r += 2
    _cabecera(ws, r, ["Fecha", "Concepto", "Monto"]); r += 1
    for ln in d["lineas"]:
        ws.cell(row=r, column=1, value=ln["fecha"].strftime("%d/%m/%Y"))
        ws.cell(row=r, column=2, value=ln["concepto"])
        _money(ws, r, 3, ln["monto"]); r += 1
    r += 1
    _total(ws, r, "Entradas", d["entradas"], 3); r += 1
    _total(ws, r, "Salidas", d["salidas"], 3); r += 1
    _total(ws, r, "Flujo neto", d["neto"], 3); r += 1
    _total(ws, r, "Saldo final", d["saldo_final"], 3)
    _anchos(ws, [14, 40, 16])
    return "flujo_efectivo"


# ── Balanza de Comprobación ───────────────────────────────────────────────────
def balanza(ws, anio, mes, periodo):
    d = posting.balanza_comprobacion(anio, mes)
    _titulo(ws, "Balanza de Comprobación", periodo)
    r = 4
    _cabecera(ws, r, ["Código", "Cuenta", "Tipo", "Debe", "Haber", "Saldo"]); r += 1
    for f in d["filas"]:
        ws.cell(row=r, column=1, value=f["codigo"])
        ws.cell(row=r, column=2, value=f["nombre"])
        ws.cell(row=r, column=3, value=f["tipo"])
        _money(ws, r, 4, f["debe"])
        _money(ws, r, 5, f["haber"])
        _money(ws, r, 6, f["saldo"])
        r += 1
    ws.cell(row=r, column=1, value="TOTALES").font = _TOTAL
    for col in range(1, 7):
        ws.cell(row=r, column=col).fill = _TOTAL_FILL
    _money(ws, r, 4, d["total_debe"], negrita=True)
    _money(ws, r, 5, d["total_haber"], negrita=True)
    r += 1
    ws.cell(row=r, column=1,
            value=("✓ Debe = Haber (cuadra)" if d["cuadra"]
                   else "⚠ Debe ≠ Haber (NO cuadra)"))
    _anchos(ws, [10, 28, 14, 14, 14, 14])
    return "balanza_comprobacion"


BUILDERS = {
    "resultados": resultados,
    "balance": balance,
    "flujo": flujo,
    "balanza": balanza,
}


def construir(reporte, anio, mes, periodo):
    """Devuelve (workbook, nombre_archivo) para el reporte y periodo dados."""
    wb = Workbook()
    ws = wb.active
    ws.title = reporte.capitalize()
    nombre = BUILDERS[reporte](ws, anio, mes, periodo)
    return wb, nombre
